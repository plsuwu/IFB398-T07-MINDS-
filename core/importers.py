import datetime
import io

import openpyxl
from django.contrib.gis.geos import Point

from core.geo_utils import parse_grid_epsg, projected_to_wgs84
from core.models import AssayResult, Drillhole, DrillholeSurvey, LithologyInterval


def _str(val):
    if val is None:
        return ""
    s = str(val).strip()
    return s if s.lower() not in ("none", "nan") else ""


def _float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _date(val):
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    return None


def _int(val):
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _build_hmap(header_row):
    """
    Build a lowercase header->column-index map from the first row of a sheet.
    Duplicate headers (e.g. two 'Mineralisation' columns) get a #2, #3 suffix
    on the later occurrences so both are reachable.
    """
    hmap = {}
    seen = {}
    for idx, val in enumerate(header_row):
        if val is None:
            continue
        key = str(val).strip().lower()
        if key in seen:
            seen[key] += 1
            hmap[f"{key}#{seen[key]}"] = idx
        else:
            seen[key] = 1
            hmap[key] = idx
    return hmap


def _h(row, hmap, key, default=None):
    """Retrieve a cell value by header name (case-insensitive)."""
    idx = hmap.get(key.lower())
    if idx is not None and idx < len(row):
        return row[idx]
    return default


def run_drillhole_import(file_obj, org, process, dry_run=False, update=False, warn=None):
    """
    Import drillhole collar, survey, lithology, and assay data from an xlsx workbook.

    file_obj  — path string, Path object, or a file-like object (e.g. Django UploadedFile)
    org       — Organisation model instance
    process   — Process model instance
    dry_run   — if True, parse and count without writing to the database
    update    — if True, overwrite existing Drillhole records matched by HOLEID
    warn      — optional callable(str) for per-row warning messages

    Returns a dict: {"counters": {...}, "errors": [...]}
    """
    if warn is None:
        warn = lambda msg: None  # noqa: E731

    # If file_obj is a file-like object that may have been partially read, reset it.
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)

    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)

    counters = {
        "collars_created": 0,
        "collars_updated": 0,
        "collars_skipped": 0,
        "collars_errors": 0,
        "surveys_created": 0,
        "surveys_errors": 0,
        "litho_created": 0,
        "litho_errors": 0,
        "assay_created": 0,
        "assay_errors": 0,
    }
    errors = []

    def _warn(msg):
        warn(msg)
        errors.append(msg)

    # ── DH Collars ────────────────────────────────────────────────────────────
    ws = wb["DH Collars"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        raise ValueError("Sheet 'DH Collars' is empty.")
    hmap = _build_hmap(header)
    drillhole_map = {}

    for row_idx, row in enumerate(rows, start=2):
        if not any(row):
            continue

        holeid = _str(_h(row, hmap, "holeid"))
        if not holeid:
            _warn(f"Collars row {row_idx}: missing HOLEID — skipped.")
            counters["collars_errors"] += 1
            continue

        grid_str  = _str(_h(row, hmap, "grid"))
        raw_east  = _float(_h(row, hmap, "mga94_east"))
        raw_north = _float(_h(row, hmap, "mga94_north"))
        collar_point = None
        source_crs   = ""

        if raw_east is not None and raw_north is not None:
            epsg = parse_grid_epsg(grid_str)
            if epsg is None:
                _warn(f"Collars row {row_idx} ({holeid}): cannot parse GRID '{grid_str}' — collar_location will be NULL.")
            else:
                try:
                    lon, lat = projected_to_wgs84(raw_east, raw_north, epsg)
                    collar_point = Point(lon, lat, srid=4326)
                    source_crs   = f"EPSG:{epsg}"
                except Exception as exc:
                    _warn(f"Collars row {row_idx} ({holeid}): CRS transform failed — {exc}.")

        drill_type_raw = _str(_h(row, hmap, "type")).upper()
        valid_types    = [c[0] for c in Drillhole.DrillType.choices]
        drill_type     = drill_type_raw if drill_type_raw in valid_types else ""

        defaults = dict(
            drill_type       = drill_type,
            company          = _str(_h(row, hmap, "company")),
            drill_company    = _str(_h(row, hmap, "drill company")),
            current_epm      = _str(_h(row, hmap, "current_epm")),
            original_epm     = _str(_h(row, hmap, "original epm")),
            year_report      = _int(_h(row, hmap, "year report")),
            company_report   = _str(_h(row, hmap, "companyreport")),
            collar_location  = collar_point,
            elevation        = _float(_h(row, hmap, "elevation")),
            depth            = _float(_h(row, hmap, "total_depth")),
            dip              = _float(_h(row, hmap, "dip")),
            azimuth          = _float(_h(row, hmap, "azimuth (tn)")),
            date_commenced   = _date(_h(row, hmap, "date commenced")),
            date_completed   = _date(_h(row, hmap, "date completed")),
            hole_id_original = _str(_h(row, hmap, "holeid_original")),
            comments         = _str(_h(row, hmap, "comments")),
            source_crs       = source_crs,
            source_easting   = raw_east,
            source_northing  = raw_north,
            organisation     = org,
            process          = process,
        )

        existing = Drillhole.objects.filter(name=holeid, organisation=org, process=process).first()
        if existing:
            if update:
                if not dry_run:
                    for field, value in defaults.items():
                        setattr(existing, field, value)
                    existing.save()
                drillhole_map[holeid] = holeid if dry_run else existing
                counters["collars_updated"] += 1
            else:
                drillhole_map[holeid] = existing
                counters["collars_skipped"] += 1
        else:
            if not dry_run:
                dh = Drillhole.objects.create(name=holeid, **defaults)
                drillhole_map[holeid] = dh
            else:
                drillhole_map[holeid] = holeid
            counters["collars_created"] += 1

    # ── DH DHSurvey ───────────────────────────────────────────────────────────
    ws   = wb["DH DHSurvey"]
    rows = ws.iter_rows(values_only=True)
    hmap = _build_hmap(next(rows, ()))

    for row_idx, row in enumerate(rows, start=2):
        if not any(row):
            continue
        holeid = _str(_h(row, hmap, "holeid"))
        dh = drillhole_map.get(holeid)
        if dh is None:
            _warn(f"Survey row {row_idx}: HOLEID '{holeid}' not found — skipped.")
            counters["surveys_errors"] += 1
            continue
        if not dry_run:
            try:
                DrillholeSurvey.objects.create(
                    drillhole   = dh,
                    depth       = _float(_h(row, hmap, "depth")) or 0.0,
                    dip         = _float(_h(row, hmap, "dip")),
                    azimuth_tn  = _float(_h(row, hmap, "azimuth (tn)")),
                    azimuth_mag = _float(_h(row, hmap, "azimuth_mag")),
                    comment     = _str(_h(row, hmap, "azi comment")),
                )
                counters["surveys_created"] += 1
            except Exception as exc:
                _warn(f"Survey row {row_idx}: {exc}")
                counters["surveys_errors"] += 1
        else:
            counters["surveys_created"] += 1

    # ── DH_Lithology ──────────────────────────────────────────────────────────
    ws   = wb["DH_Lithology"]
    rows = ws.iter_rows(values_only=True)
    hmap = _build_hmap(next(rows, ()))

    for row_idx, row in enumerate(rows, start=2):
        if not any(row):
            continue
        holeid = _str(_h(row, hmap, "holeid"))
        dh = drillhole_map.get(holeid)
        if dh is None:
            _warn(f"Lithology row {row_idx}: HOLEID '{holeid}' not found — skipped.")
            counters["litho_errors"] += 1
            continue
        if not dry_run:
            try:
                LithologyInterval.objects.create(
                    drillhole        = dh,
                    from_depth       = _float(_h(row, hmap, "from")) or 0.0,
                    to_depth         = _float(_h(row, hmap, "to")) or 0.0,
                    lithology        = _str(_h(row, hmap, "lithology")),
                    description      = _str(_h(row, hmap, "description")),
                    mineralisation   = _str(_h(row, hmap, "mineralisation")),
                    hardness         = _str(_h(row, hmap, "hardness")),
                    weathering       = _str(_h(row, hmap, "weathering")),
                    acid_reaction    = _str(_h(row, hmap, "acid")),
                    colour           = _str(_h(row, hmap, "colour")),
                    oxidation        = _str(_h(row, hmap, "oxidation")),
                    mineralisation_b = _str(_h(row, hmap, "mineralisation#2")),
                    mineralisation_2 = _str(_h(row, hmap, "mineralisation 2")),
                    alteration       = _str(_h(row, hmap, "alteration")),
                    alteration_2     = _str(_h(row, hmap, "alteration 2")),
                    veins            = _str(_h(row, hmap, "veins")),
                    recovery_pct     = _str(_h(row, hmap, "recovery %")),
                    core_size        = _str(_h(row, hmap, "core size")),
                )
                counters["litho_created"] += 1
            except Exception as exc:
                _warn(f"Lithology row {row_idx}: {exc}")
                counters["litho_errors"] += 1
        else:
            counters["litho_created"] += 1

    # ── DH_Assays ─────────────────────────────────────────────────────────────
    ws   = wb["DH_Assays"]
    rows = ws.iter_rows(values_only=True)
    hmap = _build_hmap(next(rows, ()))

    for row_idx, row in enumerate(rows, start=2):
        if not any(row):
            continue
        holeid = _str(_h(row, hmap, "holeid"))
        dh = drillhole_map.get(holeid)
        if dh is None:
            _warn(f"Assay row {row_idx}: HOLEID '{holeid}' not found — skipped.")
            counters["assay_errors"] += 1
            continue
        if not dry_run:
            try:
                AssayResult.objects.create(
                    drillhole        = dh,
                    from_depth       = _float(_h(row, hmap, "from")) or 0.0,
                    to_depth         = _float(_h(row, hmap, "to")) or 0.0,
                    lab_batch_number = _str(_h(row, hmap, "lab batch number")),
                    sample_number    = _str(_h(row, hmap, "samplenumber")),
                    comment          = _str(_h(row, hmap, "comment")),
                    au_ppm           = _float(_h(row, hmap, "au_ppm")),
                    au_ppm_check1    = _float(_h(row, hmap, "au_ppmchecks")),
                    au_ppm_check2    = _float(_h(row, hmap, "au_ppmchecks2")),
                    cu_ppm           = _float(_h(row, hmap, "cu_ppm")),
                    pb_ppm           = _float(_h(row, hmap, "pb_ppm")),
                    zn_ppm           = _float(_h(row, hmap, "zn_ppm")),
                    ag_ppm           = _float(_h(row, hmap, "ag_ppm")),
                    as_ppm           = _float(_h(row, hmap, "as_ppm")),
                    bi_ppm           = _float(_h(row, hmap, "bi_ppm")),
                    cd_ppm           = _float(_h(row, hmap, "cd_ppm")),
                    sb_ppm           = _float(_h(row, hmap, "sb_ppm")),
                    mn_ppm           = _float(_h(row, hmap, "mn_ppm")),
                    mo_ppm           = _float(_h(row, hmap, "mo_ppm")),
                    pt_ppb           = _float(_h(row, hmap, "pt_ppb")),
                    pd_ppb           = _float(_h(row, hmap, "pd_ppb")),
                    laboratory       = _str(_h(row, hmap, "laboratory")),
                    au_method        = _str(_h(row, hmap, "au_method")),
                    cu_method        = _str(_h(row, hmap, "cu_method")),
                    cu_method_2      = _str(_h(row, hmap, "cu_method2")),
                    pb_method        = _str(_h(row, hmap, "pb_method")),
                    zn_method        = _str(_h(row, hmap, "zn_method")),
                    ag_method        = _str(_h(row, hmap, "ag_method")),
                    as_method        = _str(_h(row, hmap, "as_method")),
                    bi_method        = _str(_h(row, hmap, "bi_method")),
                    cd_method        = _str(_h(row, hmap, "cd_method")),
                    sb_method        = _str(_h(row, hmap, "sb_method")),
                    mn_method        = _str(_h(row, hmap, "mn_method")),
                    mo_method        = _str(_h(row, hmap, "mo_method")),
                    pt_method        = _str(_h(row, hmap, "pt_method")),
                    pd_method        = _str(_h(row, hmap, "pd_method")),
                )
                counters["assay_created"] += 1
            except Exception as exc:
                _warn(f"Assay row {row_idx}: {exc}")
                counters["assay_errors"] += 1
        else:
            counters["assay_created"] += 1

    return {"counters": counters, "errors": errors}
